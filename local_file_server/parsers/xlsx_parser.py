MAX_TEXT_BYTES = 50 * 1024


def parse_xlsx(file_path: str) -> tuple[str, bool]:
    """Extract text from an XLSX file as CSV-like format."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        return "[Error: openpyxl not installed. Run: pip install openpyxl]", False

    wb = load_workbook(file_path, read_only=True, data_only=True)
    lines = []
    total_len = 0
    truncated = False

    for sheet in wb.sheetnames:
        ws = wb[sheet]
        lines.append(f"--- Sheet: {sheet} ---")
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            line = " | ".join(cells)
            if total_len + len(line) > MAX_TEXT_BYTES:
                truncated = True
                break
            lines.append(line)
            total_len += len(line)
        if truncated:
            break

    wb.close()
    return "\n".join(lines), truncated
